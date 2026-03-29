import json
import os
import zipfile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage

from ..config import STATIC_DIR, EXPORTS_DIR
from ..database import get_db


async def _get_project_and_pages(project_id: int):
    db = await get_db()
    try:
        cursor = await db.execute("SELECT * FROM projects WHERE id = ?", (project_id,))
        project = await cursor.fetchone()
        if not project:
            raise ValueError("Project not found")
        project = {k: project[k] for k in project.keys()}

        cursor = await db.execute(
            "SELECT * FROM pages WHERE project_id = ? ORDER BY page_number",
            (project_id,),
        )
        pages = [{k: r[k] for k in r.keys()} for r in await cursor.fetchall()]
        return project, pages
    finally:
        await db.close()


def _resolve_image_path(image_url: str) -> str:
    """Convert URL path like /static/uploads/1/file.png to absolute filesystem path."""
    return str(STATIC_DIR / image_url.lstrip("/").removeprefix("static/"))


async def export_zip(project_id: int) -> str:
    project, pages = await _get_project_and_pages(project_id)

    export_dir = EXPORTS_DIR / str(project_id)
    os.makedirs(export_dir, exist_ok=True)
    zip_path = export_dir / f"{project['child_name']}_book.zip"

    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        meta = {
            "child_name": project["child_name"],
            "story_type": project["story_type"],
            "pages": [],
        }

        for page in pages:
            page_info = {
                "page_number": page["page_number"],
                "page_type": page["page_type"],
                "text": page["text"],
            }
            meta["pages"].append(page_info)

            if page.get("current_image_path"):
                local_path = _resolve_image_path(page["current_image_path"])
                if os.path.exists(local_path):
                    arcname = f"images/page_{page['page_number']:02d}.png"
                    zf.write(local_path, arcname)

        zf.writestr("metadata.json", json.dumps(meta, ensure_ascii=False, indent=2))

    db = await get_db()
    try:
        file_url = f"/static/exports/{project_id}/{project['child_name']}_book.zip"
        await db.execute(
            "INSERT INTO exports (project_id, format, file_path) VALUES (?, ?, ?)",
            (project_id, "zip", file_url),
        )
        await db.commit()
    finally:
        await db.close()

    return file_url


async def export_excel(project_id: int) -> str:
    project, pages = await _get_project_and_pages(project_id)

    export_dir = EXPORTS_DIR / str(project_id)
    os.makedirs(export_dir, exist_ok=True)
    xlsx_path = export_dir / f"{project['child_name']}_book.xlsx"

    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Projekt"
    info_rows = [
        ("Imię", project["child_name"]),
        ("Wiek", project["child_age"]),
        ("Płeć", project["child_gender"]),
        ("Motyw", project["story_type"]),
        ("Hobby", project["hobby"]),
        ("Przesłanie", project["moral"]),
        ("Status", project["status"]),
    ]
    for r, (label, val) in enumerate(info_rows, 1):
        ws_info.cell(row=r, column=1, value=label)
        ws_info.cell(row=r, column=2, value=str(val))

    ws_pages = wb.create_sheet("Strony")
    ws_pages.append(["Nr", "Typ", "Tekst", "Prompt obrazkowy"])

    for page in pages:
        row_idx = ws_pages.max_row + 1
        ws_pages.cell(row=row_idx, column=1, value=page["page_number"])
        ws_pages.cell(row=row_idx, column=2, value=page["page_type"])
        ws_pages.cell(row=row_idx, column=3, value=page.get("text", ""))
        ws_pages.cell(row=row_idx, column=4, value=page.get("image_prompt", ""))

        if page.get("current_image_path"):
            local_path = _resolve_image_path(page["current_image_path"])
            if os.path.exists(local_path):
                try:
                    img = XlImage(local_path)
                    img.width = 150
                    img.height = 150
                    ws_pages.add_image(img, f"E{row_idx}")
                except Exception:
                    pass

    ws_pages.column_dimensions["C"].width = 60
    ws_pages.column_dimensions["D"].width = 60

    wb.save(str(xlsx_path))

    db = await get_db()
    try:
        file_url = f"/static/exports/{project_id}/{project['child_name']}_book.xlsx"
        await db.execute(
            "INSERT INTO exports (project_id, format, file_path) VALUES (?, ?, ?)",
            (project_id, "excel", file_url),
        )
        await db.commit()
    finally:
        await db.close()

    return file_url
